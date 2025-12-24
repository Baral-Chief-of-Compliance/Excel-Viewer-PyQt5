from typing import Tuple
from datetime import date, datetime, time

import openpyxl



NUMBER_TYPE = 'Числовой'
TEXT_TYPE = 'Текстовый'
DATA_TIME_TYPE = 'Дата/время'
OTHER_TYPE = 'Смешанный'


BASE_TYPE_FILTER_STRUCT = {
    NUMBER_TYPE : {
        'count': 0,
        'min' : 0,
        'max' : 0,
    },
    TEXT_TYPE: 0,
    DATA_TIME_TYPE : 0,
    OTHER_TYPE: 0
}


class ExcelConverter:
    """Класс для работы с excle файлами"""
    wb = None


    def __init__(self, filepath : str) -> None:
        self.filepath = filepath
        self._open_file(filepath)


    def _open_file(self, filepath) -> None:
        """Открыть excel файл"""
        self.wb = openpyxl.load_workbook(filepath)

    
    def get_work_book(self) -> openpyxl.Workbook:
        """Получить данные в массиве"""
        return self.wb


    def get_filename(self) -> str:
        """Получить наименование файла"""
        filepath_split = self.filepath.split('/')
        return filepath_split[len(filepath_split) - 1]
    

    def get_cols_number(self) -> int:
        """Получить кол-во столбцов"""
        return self.wb.worksheets[0].max_column
    

    def get_rows_number(self) -> int:
        """Получить кол-во строк"""
        return self.wb.worksheets[0].max_row


    def get_info_about_excel_file(self) -> dict:
        """Получить информацию об excel файле"""
        cols_num = self.get_cols_number()
        rows_num = self.get_rows_number()

        return {
            'filename': self.get_filename(),
            'sheetname': self.wb.worksheets[0].title,
            'cols_num': cols_num,
            'rows_num': rows_num
        }
    
    def get_table_headers(self) -> list:
        """Получить заголовки таблицы, если такие есть
        если нет list будет пустым"""

        headers = [cell.value for cell in self.wb.worksheets[0][1]]

        return headers
    

    def count_empty_values_in_col(self, col) -> int:
        """Насчитать кол-во пустых значений"""
        empty_count = 0
        for cell in col:
            if cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == ''):
                    empty_count += 1
        
        return empty_count
    

    def define_type_in_col(self, col) -> str:
        """Определить тип столбца в таблице"""
        col_info = {}

        for cell in col:
            if cell.value is not None:
                if isinstance(cell.value, str):
                    col_info[TEXT_TYPE] = 0

                elif isinstance(cell.value, (int, float)):
                    col_info[NUMBER_TYPE] = 0

                elif isinstance(cell.value, (time, date, datetime)):
                    col_info[DATA_TIME_TYPE] = 0
                else:
                    col_info[OTHER_TYPE] = 0


        if len(col_info) == 1:
            return list(col_info.keys())[0]
        else:
            return OTHER_TYPE
        
    def find_mix_and_max_number_in_col(self, col) -> Tuple[int, int]:
        """Получить максимальное и минимальное числовое значение в столбце
        Возвращает сначала минимальное, потом максмимальное
        """
        values : list = []
        for cell in col:
            if isinstance(cell.value, (int, float)):
                values.append(cell.value)

        return min(values), max(values)
    

    def analys_col(self, col) -> dict:
        """Проанализировать столбец"""
        col_type = self.define_type_in_col(col)
        analys = {
            'name': 'test',
            'type': col_type,
            'empty': self.count_empty_values_in_col(col),
        }

        if col_type == NUMBER_TYPE:
            analys['max'], analys['min'] = self.find_mix_and_max_number_in_col(col)
            analys['mean'] = ( analys['max'], analys['min']) / 2
        else:
            analys['max'] = None
            analys['min'] = None
            analys['mean'] = None

        return analys
        
    

    def data_analysis(self):
        """Анализ данных"""

        cols = self.wb.worksheets[0].iter_cols(
                min_col=1,
                max_col=self.wb.worksheets[0].max_column
            )
        
        cols_analys = []
        for col in cols:
            cols_analys.append(self.analys_col(col))
        
        return {
            'brief' : self.get_info_about_excel_file(),
            'headers': self.get_table_headers(),
            'cols': cols_analys
        }
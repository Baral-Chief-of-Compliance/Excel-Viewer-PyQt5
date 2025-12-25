from datetime import datetime

from PyQt5.QtWidgets import QLabel, QPushButton, QVBoxLayout,\
QHBoxLayout, QWidget, QFileDialog, QMessageBox, QTableWidget,\
QTableWidgetItem
from PyQt5.QtCore import pyqtSignal
from PyQt5 import QtWidgets
from openpyxl import Workbook

from utils.excel_converters import ExcelConverter
from .styles import QLABEL_INFO_EXCEL_STYLE, QLABEL_BRIED_INFO_TABLE


class ExcelLoadDialog(QWidget):

    excel_loaded = pyqtSignal(object)
    excel_info_loaded = pyqtSignal(dict)
    

    def __init__(self):
        super(ExcelLoadDialog, self).__init__()
        self.select_excel_button : QPushButton = None
        self.excel_label : QLabel = None
        self.clear_button : QPushButton = None 

        self.excel_info_layout : QHBoxLayout = QHBoxLayout()
        self.excel_file = None
        self.workbook = None

        self.init_ui()


    def init_select_excel_btn(self):
        """Настроить кнопки для выбора excel файла"""
        self.select_excel_button = QPushButton("Выбрать Excel файл")
        self.select_excel_button.clicked.connect(self.select_excel_file)
        self.select_excel_button.setMinimumHeight(40)


    def init_excel_label(self):
        """Настроить отображение информации о excel файле"""
        self.excel_label = QLabel("Файл не выбран")
        self.excel_label.setWordWrap(True)


    def set_excel_label(self, filepath : str):
        """Сброс текст на кнопке с excel файлом"""
        self.excel_label.setText(filepath)


    def init_clear_button(self):
        """Настроить кнопку открепления файла"""
        self.clear_button = QPushButton("x")
        self.clear_button.setFixedSize(25, 25)
        self.clear_button.setToolTip("Очистить")
        self.clear_button.clicked.connect(self.clear_excel_file)
        self.clear_button.setEnabled(False)


    def enable_claer_button(self):
        """Сделать активным кнопку открепления"""
        self.clear_button.setEnabled(True)


    def disable_claer_button(self):
        """Сдеалать неактивным кнопку открепления"""
        self.clear_button.setEnabled(False)

        
    def init_ui(self):
        """Инициализируем UI"""
        layout = QVBoxLayout()
        layout.setSpacing(10)

        # Кнопка для загрузки
        self.init_select_excel_btn()

        layout.addWidget(self.select_excel_button)

        #Qtlabel для отображения информации о файле
        self.init_excel_label()

        #кнопка для открепления файла
        self.init_clear_button()

        self.excel_info_layout.addWidget(self.excel_label, 1)
        self.excel_info_layout.addWidget(self.clear_button)

        layout.addLayout(self.excel_info_layout)

        self.setLayout(layout)


    def select_excel_file(self):
        """Выбрать excel файл """
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Выберите Excel файл",
                "",
                "Excel Files (*.xlsx *.xls);;All Files (*)"
            )

            if file_path:
                if not (file_path.endswith('.xlsx') or file_path.endswith('.xls')):
                    QMessageBox.warning(self, "Предупреждение", "Выберите файл с расширением .xlsx или .xls")
                    return
                
                #логика обработки файла
                ec = ExcelConverter(filepath=file_path)
                
                self.workbook = ec.get_work_book()
                self.excel_info = ec.data_analysis()

                self.set_excel_label(file_path)
                self.enable_claer_button()


                self.excel_loaded.emit(self.workbook)
                self.excel_info_loaded.emit(self.excel_info)


        except Exception as ex:
            QMessageBox.critical(self, "Ошибка", "Не удалось открыть файл: {}".format(ex))


    def clear_excel_file(self):
        """Открепить файл"""
        self.set_excel_label("Файл не выбран")
        self.disable_claer_button()
        self.workbook = None
        self.excel_file = {}

        self.excel_loaded.emit(None)
        self.excel_info_loaded.emit({})


class ExcelTableWidget(QTableWidget):

    wb : Workbook = None

    def __init__(self, wb : Workbook):
        super(ExcelTableWidget, self).__init__()
        self.wb = wb
        self.init_table()
        self.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)


    def init_table(self):
        """Инициализация таблицы"""
        if self.wb is not None:
            cols = self.wb.worksheets[0].iter_cols(
                min_col=1,
                max_col=self.wb.worksheets[0].max_column
            )

            rows = self.wb.worksheets[0].iter_rows(
                min_row=1,
                max_row=self.wb.worksheets[0].max_row

            )

            self.setColumnCount(self.wb.worksheets[0].max_column)

            self.setRowCount(self.wb.worksheets[0].max_row)


            for row in rows:
                for cell in row:
                    value = None
                    if isinstance(cell.value, datetime):
                        value = cell.value.strftime("%d.%m.%Y %H:%M")
                    else:
                        value = cell.value
                    item = QTableWidgetItem(value)

                    self.setItem(
                        cell.row,
                        cell.column,
                        item
                        
                    )

            self.resizeColumnsToContents()


class ExcelInfoWidget(QWidget):
    info : dict = None
    KEYS = [
        'brief',
        'headers',
        'cols',
    ]

    def __init__(self, info: dict):
        super(ExcelInfoWidget, self).__init__()
        
        self.init_status = True

        for k in self.KEYS:
            if k not in info:
                self.init_status = False

        if self.init_status:
            self.info = info
            self.title = QLabel("Информация о таблице")
            self.layout : QVBoxLayout = QVBoxLayout()
            self.excel_cols_info = QTableWidget()
            self.brief_info : QWidget = QWidget()
            self.init_ui()


    def init_titel(self):
        """Настроить заголовок для блока c информацией о таблице"""
        self.title.setStyleSheet(QLABEL_INFO_EXCEL_STYLE)

    
    def get_label_for_brief(self, lable : str) -> QLabel:
        """Получить QLabel для краткой информации о таблице"""
        return QLabel(lable).setStyleSheet(QLABEL_BRIED_INFO_TABLE)


    def init_brief_info_table(self):
        """Настроить краткую информацию о таблице (колонки, строки, список имен)"""

        info_filename = QHBoxLayout()
        info_filename.addWidget(self.get_label_for_brief('Наименование файла'))
        info_filename.addWidget(self.get_label_for_brief(str(self.info['brief']['filename'])))
        self.brief_info.(info_filename)

        info_sheetname = QHBoxLayout()
        info_sheetname.addWidget(self.get_label_for_brief('Наименование листа'))
        info_sheetname.addWidget(self.get_label_for_brief(str(self.info['brief']['sheetname'])))
        self.brief_info.addWidget(info_sheetname)

        info_row = QHBoxLayout()
        info_row.addWidget(self.get_label_for_brief('Количество строк'))
        info_row.addWidget(self.get_label_for_brief(str(self.info['brief']['rows_num'])))
        self.brief_info.addWidget(info_row)

        info_col = QHBoxLayout()
        info_col.addWidget(self.get_label_for_brief('Количество столбцов'))
        info_col.addWidget(self.get_label_for_brief(str(self.info['brief']['cols_num'])))
        self.brief_info.addWidget(info_col)


    
    def init_info_table(self):
        """Настроить таблицу с информацией о каждой колонке"""
        self.excel_cols_info.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)

        self.excel_cols_info.setColumnCount(6)
        self.excel_cols_info.setRowCount(len(self.info['cols']) + 1)

        #Иницилизируем heade таблицы
        self.excel_cols_info.setItem(0,0,QTableWidgetItem('Столбец'))
        self.excel_cols_info.setItem(0,1,QTableWidgetItem('Тип'))
        self.excel_cols_info.setItem(0,2,QTableWidgetItem('Пустых'))
        self.excel_cols_info.setItem(0,3,QTableWidgetItem('Min'))
        self.excel_cols_info.setItem(0,4,QTableWidgetItem('Max'))
        self.excel_cols_info.setItem(0,5,QTableWidgetItem('Mean'))


        for index_row, ic in enumerate(self.info['cols']):
            for index_col, k in enumerate(ic.keys()):
                self.excel_cols_info.setItem(
                    index_row + 1,
                    index_col,
                    QTableWidgetItem(ic[k])
                )

        self.excel_cols_info.resizeColumnsToContents()


    def init_ui(self):
        self.layout.setSpacing(10)

        self.init_titel()
        self.init_info_table()
        self.init_brief_info_table()

        self.layout.addWidget(self.title)
        
        self.layout.addStretch(0)

        self.layout.addWidget(self.excel_cols_info)

        self.layout.addStretch(0)

        self.layout.addWidget(self.brief_info)

        self.setLayout(self.layout)
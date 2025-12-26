from datetime import datetime

from openpyxl import Workbook
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem


class ExcelTableWidget(QTableWidget):

    wb : Workbook = None

    def __init__(self, wb : Workbook):
        super(ExcelTableWidget, self).__init__()
        self.wb = wb
        self.init_table()
        self.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)


    def init_table(self):
        """Инициализация таблицы"""
        if self.wb is not None:
            cols = self.wb.worksheets[0].iter_cols(
                min_col=1,
                max_col=self.wb.worksheets[0].max_column
            )

            rows = self.wb.worksheets[0].iter_rows(
                min_row=1,
                max_row=self.wb.worksheets[0].max_row

            )

            self.setColumnCount(self.wb.worksheets[0].max_column)

            self.setRowCount(self.wb.worksheets[0].max_row)


            for row in rows:
                for cell in row:
                    value = None
                    if isinstance(cell.value, datetime):
                        value = cell.value.strftime("%d.%m.%Y %H:%M")
                    else:
                        value = cell.value
                    item = QTableWidgetItem(value)

                    self.setItem(
                        cell.row,
                        cell.column,
                        item
                        
                    )

            self.resizeColumnsToContents()